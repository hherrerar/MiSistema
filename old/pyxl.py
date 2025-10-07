from openpyxl import Workbook, load_workbook

wb = load_workbook('grades.xlsx') 

ws = wb.active
ws['A2'].value="Test"

wb.save('grades.xlsx')

print(ws['A2'].value)
